import numpy as np
import pandas as pd
from openpyxl import load_workbook

main_path = "/Users/darkesthj/Library/CloudStorage/OneDrive-Personal/Documentos/Main.xlsx"
test_path = "/Users/darkesthj/Library/CloudStorage/OneDrive-Personal/Documentos/test.xlsx"

df_raw = pd.read_excel(main_path, 'DataRaw') # DataFrame de Main
main_loaded = load_workbook(main_path) # Todo el file de Main (1)

def update_dataPy():
    count_of_rows = 0 # Esto ayuda a definir la ultima fila que se copiara

    for i in range(len(df_raw['Income'])):
        if df_raw['Income'][i] != 0:
            count_of_rows += 1

    # Eliminar 'DataPy' existente en Main
    try:
        main_loaded.remove(main_loaded["DataPy"])
        main_loaded.save(main_path)
    except:
        print('No DataPy found in Main.xlsx')

    # Crear nuevo DataPy dentro del test.xlsx
    my_new_data = df_raw[:][:count_of_rows]
    my_new_data.to_excel(test_path, sheet_name = 'DataPy', index=False)

    # Cargar main y test actualizados
    main_loaded = load_workbook(main_path) # Todo el file de Main (2)
    test_loaded = load_workbook(test_path) # Todo el file de Test

    # Agregar el test.xlsx a Main.xlsx
    tab_or = test_loaded["DataPy"]
    tab_des = main_loaded.create_sheet(title="DataPy")

    for row in tab_or.iter_rows():
        for celda in row:
            tab_des[celda.coordinate].value = celda.value

    main_loaded.save(main_path) # Save!

update_dataPy()
